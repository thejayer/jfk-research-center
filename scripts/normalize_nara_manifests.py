#!/usr/bin/env python3
"""
scripts/normalize_nara_manifests.py

Normalize the NARA JFK Records Collection XLSX manifests into one unified
CSV suitable for loading into `jfk_raw.nara_manifest`.

Schema drift across releases is absorbed here:
  - 2017-2018, 2023: [File Name, Record Num, ..., Agency, Doc Date, Doc Type, ...]
  - 2022:            same as above but no Agency column
  - 2021:            columns renamed (Record Number, File Title, To, From, ...)

Unified output columns match `jfk_raw.nara_manifest`:
  record_num, file_name, release_date, formerly_withheld, agency,
  doc_date, doc_type, file_num, to_name, from_name, title, num_pages,
  originator, record_series, review_date, comments, pages_released,
  release_set, agency_from_prefix, pdf_url

Usage:
  python3 scripts/normalize_nara_manifests.py \
    --inputs /tmp/jfk-2017-2018.xlsx /tmp/jfk-2021.xlsx /tmp/jfk-2022.xlsx /tmp/jfk-2023.xlsx \
    --out    /tmp/jfk-data/jfk-records.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


# Source-column → canonical-column aliases
ALIAS = {
    "record_num":        ["Record Num", "Record Number"],
    "file_name":         ["File Name", "File Title"],
    "release_date":      ["NARA Release Date"],
    "formerly_withheld": ["Formerly Withheld"],
    "agency":            ["Agency"],
    "doc_date":          ["Doc Date", "Document Date"],
    "doc_type":          ["Doc Type", "Document Type"],
    "file_num":          ["File Num", "File Number"],
    "to_name":           ["To Name", "To"],
    "from_name":         ["From Name", "From"],
    "title":             ["Title"],
    "num_pages":         ["Num Pages", "Original Document Pages"],
    "originator":        ["Originator"],
    "record_series":     ["Record Series"],
    "review_date":       ["Review Date"],
    "comments":          ["Comments"],
    "pages_released":    ["Pages Released", "Document Pages in PDF"],
}
OUT_COLS = list(ALIAS.keys()) + ["release_set", "agency_from_prefix", "pdf_url"]

# RIF prefix → source agency (when the explicit Agency column is missing)
PREFIX_AGENCY = {
    "104": "CIA",
    "124": "FBI",
    "144": "Senate (Church Committee)",
    "157": "Senate",
    "176": "NSA",
    "178": "Department of State",
    "179": "Department of Justice",
    "180": "HSCA",
    "194": "Department of Defense",
    "198": "U.S. Marine Corps",
    "202": "U.S. Secret Service",
    "206": "Warren Commission",
}

# archives.gov roots per release
PDF_ROOT = {
    "2017-2018": "https://www.archives.gov/files/research/jfk/releases",
    "2021":      "https://www.archives.gov/files/research/jfk/releases/2021",
    "2022":      "https://www.archives.gov/files/research/jfk/releases/2022",
    "2023":      "https://www.archives.gov/files/research/jfk/releases/2023",
}

RELEASE_FROM_FILENAME = re.compile(r"jfk[-_]?(\d{4}(?:-\d{4})?)")


def release_label_from_path(path: Path) -> str:
    """Infer 2017-2018 / 2021 / 2022 / 2023 from the input filename."""
    stem = path.stem.lower()
    m = RELEASE_FROM_FILENAME.search(stem)
    if m:
        return m.group(1)
    raise SystemExit(f"[error] cannot infer release from filename: {path}")


def clean(v: object) -> str:
    return "" if v is None else str(v).strip()


def normalize_row(header: tuple, row: tuple, release_set: str) -> dict:
    h2i: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        h2i[h.strip()] = i

    def get(field: str) -> str:
        for src in ALIAS[field]:
            if src in h2i:
                return clean(row[h2i[src]])
        return ""

    rec = {k: get(k) for k in ALIAS}
    rec["release_set"] = release_set

    rn = rec["record_num"]
    prefix = rn.split("-")[0] if rn else ""
    rec["agency_from_prefix"] = PREFIX_AGENCY.get(prefix, "")

    fn = rec["file_name"].strip()
    if fn and not fn.lower().endswith(".pdf"):
        fn = f"{rn}.pdf" if rn else ""
    rec["pdf_url"] = f"{PDF_ROOT[release_set]}/{fn}" if fn and release_set in PDF_ROOT else ""
    return rec


def process_file(path: Path, writer: csv.writer) -> int:
    release_set = release_label_from_path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    count = 0
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = normalize_row(header, row, release_set)
        if not rec["record_num"]:
            continue
        writer.writerow([rec[c] for c in OUT_COLS])
        count += 1
    return count


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n", 1)[0])
    ap.add_argument("--inputs", nargs="+", type=Path, required=True,
                    help="One or more XLSX manifests")
    ap.add_argument("--out",    type=Path, required=True,
                    help="Output CSV path")
    args = ap.parse_args()

    args.out.parent.mkdir(parents=True, exist_ok=True)

    with args.out.open("w", newline="") as fd:
        w = csv.writer(fd, quoting=csv.QUOTE_ALL)
        w.writerow(OUT_COLS)
        total = 0
        for path in args.inputs:
            n = process_file(path, w)
            print(f"{path.name}: {n} rows", file=sys.stderr)
            total += n
        print(f"TOTAL: {total} rows → {args.out}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
